"""# Lista

"""**FBREF** - 2"""

from openpyxl import load_workbook

# Carregar o arquivo Excel
filename = input("Por favor, faça o upload do arquivo Excel: ")
wb = load_workbook(filename)
ws = wb.active

# Iterar sobre as células na coluna A (ignorando o cabeçalho)
for row, cell in enumerate(ws['A']):
    if row == 0:  # Ignorar o cabeçalho
        continue

    # Verificar se a célula tem um hyperlink
    if cell.hyperlink:
        url = cell.hyperlink.target

        # Dividir a URL para obter o ID e o histórico
        parts = url.split('/')
        player_id = parts[-2]
        player_name = parts[-1]

        # Escrever os resultados nas colunas B e C
        ws.cell(row=row+1, column=2, value=player_id)
        ws.cell(row=row+1, column=3, value=player_name)

# Salvar o arquivo Excel modificado
wb.save(filename)
